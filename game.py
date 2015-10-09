import cmd
import textwrap
import room
from openpyxl import load_workbook

#https://docs.google.com/spreadsheets/d/1Z24N7mSTCwnemRWJzFwy86Fepv6Nw0xrH5ky4x0bY04/edit?usp=sharing
#name	description	Nvalue	Svalue	Evalue	Wvalue

#https://docs.google.com/spreadsheets/d/1oLr3DZA4KWR8YrN73oH6NXhjB97LGkdMiy_m7TG42pI/edit?usp=sharing

class Game(cmd.Cmd):
	def __init__(self):
		cmd.Cmd.__init__(self)
		
		#doc = gc.open_by_url('https://docs.google.com/spreadsheets/d/1Z24N7mSTCwnemRWJzFwy86Fepv6Nw0xrH5ky4x0bY04/edit?usp=sharing')
		
		excel_load = load_workbook('WordGamePyData.xlsx')
		
		sheet_ranges = excel_load['Sheet1']
		print(sheet_ranges['A2'].value)

		print excel_load.get_sheet_names()
		#['Sheet2', 'New Title', 'Sheet1']
		
		self.loc = room.Room(id=1, name=sheet_ranges['A2'].value, description=sheet_ranges['B2'].value, neighbours={'n':sheet_ranges['C2'].value,
																													's':sheet_ranges['D2'].value,
																													'e':sheet_ranges['E2'].value,
																													'w':sheet_ranges['F2'].value})
		self.look()
		
	def move(self, dir):
		newroom = self.loc._neighbour(dir)
		if newroom is None:
			print("Hummm that is not the way!")
		else:
			excel_load = load_workbook('WordGamePyData.xlsx')
			sheet_ranges = excel_load['Sheet1']
			cellA = "A"+str(int(newroom)+1)
			cellB = "B"+str(int(newroom)+1)
			cellC = "C"+str(int(newroom)+1)
			cellD = "D"+str(int(newroom)+1)
			cellE = "E"+str(int(newroom)+1)
			cellF = "F"+str(int(newroom)+1)
			self.loc = room.Room(id=newroom, name=sheet_ranges[cellA].value, description=sheet_ranges[cellB].value, neighbours={'n':sheet_ranges[cellC].value,
																																's':sheet_ranges[cellD].value,
																																'e':sheet_ranges[cellE].value,
																																'w':sheet_ranges[cellF].value})
			self.look()
	
	def look(self):
		print(self.loc.name)
		print("")
		for line in textwrap.wrap(self.loc.description, 70):
			print(line)
		
	def do_n(self, args):
		"""Go north"""
		self.move('n')
		
	def do_s(self, args):
		"""Go south"""
		self.move('s')
		
	def do_e(self, args):
		"""Go east"""
		self.move('e')

	def do_w(self, args):
		"""Go west"""
		self.move('w')
		
	def do_quit(self, args):
		print("End of game")
		return True
		
if __name__ == "__main__":
	Game().cmdloop()